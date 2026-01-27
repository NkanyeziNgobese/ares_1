"""Inspect the Volve production workbook without heavy processing."""

from __future__ import annotations

import json
import os
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
DEFAULT_XLSX = ROOT / "data" / "volve" / "production" / "volve_production_data.xlsx"


def resolve_input_path() -> Path:
    env_path = os.getenv("VOLVE_PRODUCTION_XLSX")
    if env_path:
        candidate = Path(env_path)
        if not candidate.is_absolute():
            candidate = (ROOT / candidate).resolve()
        return candidate
    return DEFAULT_XLSX


def main() -> int:
    xlsx_path = resolve_input_path()
    if not xlsx_path.exists():
        print(f"Error: Volve production XLSX not found at {xlsx_path}", file=sys.stderr)
        return 1

    xls = pd.ExcelFile(xlsx_path, engine="openpyxl")
    print("Sheets:", ", ".join(xls.sheet_names))

    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    schema = {}

    for sheet in xls.sheet_names:
        # Read a sample to keep this lightweight while still inferring dtypes.
        df = pd.read_excel(xlsx_path, sheet_name=sheet, nrows=1000, engine="openpyxl")
        print(f"\nSheet: {sheet}")
        print(df.head(5).to_string(index=False))

        ws = workbook[sheet]
        row_count = max(ws.max_row - 1, 0)
        schema[sheet] = {
            "columns": [str(col) for col in df.columns.tolist()],
            "dtypes": {str(col): str(dtype) for col, dtype in df.dtypes.items()},
            "row_count": row_count,
        }

    workbook.close()

    outputs_dir = ROOT / "outputs"
    outputs_dir.mkdir(parents=True, exist_ok=True)
    schema_path = outputs_dir / "volve_production_schema.json"
    schema_path.write_text(json.dumps(schema, indent=2))
    print(f"\nWrote schema summary to {schema_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
